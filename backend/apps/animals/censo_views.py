"""
Census views for the Animales module — full tenant-level animal listing with
column picker, filters, pagination, and PDF/Excel export.
"""
import io

from django.db.models import Q
from django.http import HttpResponse
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import IsGestion
from .models import Animal

# ── Column definitions ────────────────────────────────────────────────────────

COLUMN_DEFS = [
    {"key": "numero_anilla",        "label": "Nº Anilla"},
    {"key": "fecha_nacimiento",     "label": "Fecha nacimiento"},
    {"key": "sexo",                 "label": "Sexo"},
    {"key": "variedad",             "label": "Variedad"},
    {"key": "estado",               "label": "Estado"},
    {"key": "socio_nombre",         "label": "Propietario"},
    {"key": "ganaderia_nacimiento", "label": "Ganadería nacimiento"},
    {"key": "ganaderia_actual",     "label": "Ganadería actual"},
    {"key": "padre_anilla",         "label": "Padre (anilla)"},
    {"key": "madre_anilla",         "label": "Madre (anilla)"},
    {"key": "fecha_baja",           "label": "Fecha baja"},
    {"key": "motivo_baja",          "label": "Motivo baja"},
    {"key": "fecha_incubacion",     "label": "Fecha incubación"},
]

_ESTADOS_ACTIVOS = ["REGISTRADO", "MODIFICADO", "APROBADO", "EVALUADO"]
_ESTADOS_NO_ACTIVOS = ["BAJA", "RECHAZADO", "SOCIO_EN_BAJA"]

DEFAULT_COLUMNS = [
    "numero_anilla", "fecha_nacimiento", "sexo", "variedad",
    "estado", "socio_nombre",
]

SORTABLE_FIELD_MAP = {
    "numero_anilla":        "numero_anilla",
    "fecha_nacimiento":     "fecha_nacimiento",
    "sexo":                 "sexo",
    "variedad":             "variedad",
    "estado":               "estado",
    "socio_nombre":         "socio__nombre_razon_social",
    "ganaderia_nacimiento": "ganaderia_nacimiento",
    "ganaderia_actual":     "ganaderia_actual",
    "fecha_baja":           "fecha_baja",
    "fecha_incubacion":     "fecha_incubacion",
}


def _serialize_animal(animal):
    madre_anilla = None
    if animal.madre_animal_id and animal.madre_animal:
        madre_anilla = animal.madre_animal.numero_anilla
    elif animal.madre_lote_id and animal.madre_lote:
        madre_anilla = animal.madre_lote.nombre
    elif animal.madre_lote_externo:
        madre_anilla = animal.madre_lote_externo

    def _fmt_date(d):
        return d.strftime("%d-%m-%Y") if d else ""

    return {
        "id":                    str(animal.id),
        "socio_id":              str(animal.socio_id) if animal.socio_id else "",
        "numero_anilla":         animal.numero_anilla,
        "fecha_nacimiento":      _fmt_date(animal.fecha_nacimiento),
        "sexo":                  animal.get_sexo_display(),
        "variedad":              animal.get_variedad_display(),
        "estado":                animal.estado,
        "socio_nombre":          animal.socio.nombre_razon_social if animal.socio_id else "",
        "ganaderia_nacimiento":  animal.ganaderia_nacimiento,
        "ganaderia_actual":      animal.ganaderia_actual,
        "padre_anilla":          animal.padre.numero_anilla if animal.padre_id and animal.padre else "",
        "madre_anilla":          madre_anilla or "",
        "fecha_baja":            _fmt_date(animal.fecha_baja),
        "motivo_baja":           animal.motivo_baja.nombre if animal.motivo_baja_id and animal.motivo_baja else "",
        "fecha_incubacion":      _fmt_date(animal.fecha_incubacion),
    }


def _build_queryset(tenant, params):
    qs = Animal.all_objects.filter(tenant=tenant).select_related(
        "socio", "padre", "madre_animal", "madre_lote", "motivo_baja"
    )

    # Filters
    search = params.get("search", "").strip()
    if search:
        qs = qs.filter(
            Q(numero_anilla__icontains=search) |
            Q(socio__nombre_razon_social__icontains=search) |
            Q(ganaderia_nacimiento__icontains=search) |
            Q(ganaderia_actual__icontains=search)
        )

    activo = params.get("activo")
    if activo == "true":
        qs = qs.filter(estado__in=_ESTADOS_ACTIVOS)
    elif activo == "false":
        qs = qs.filter(estado__in=_ESTADOS_NO_ACTIVOS)

    if variedad := params.get("variedad"):
        qs = qs.filter(variedad=variedad)

    if estado := params.get("estado"):
        qs = qs.filter(estado=estado)

    if sexo := params.get("sexo"):
        qs = qs.filter(sexo=sexo)

    propietario = params.get("propietario")
    if propietario == "con":
        qs = qs.filter(socio__isnull=False)
    elif propietario == "sin":
        qs = qs.filter(socio__isnull=True)

    fecha_desde = params.get("fecha_desde")
    if fecha_desde:
        qs = qs.filter(fecha_nacimiento__gte=fecha_desde)

    fecha_hasta = params.get("fecha_hasta")
    if fecha_hasta:
        qs = qs.filter(fecha_nacimiento__lte=fecha_hasta)

    # Ordering
    order_by = params.get("order_by", "numero_anilla")
    order_dir = params.get("order_dir", "asc")
    db_field = SORTABLE_FIELD_MAP.get(order_by, "numero_anilla")
    if order_dir == "desc":
        db_field = f"-{db_field}"
    qs = qs.order_by(db_field)

    return qs


# ── API views ─────────────────────────────────────────────────────────────────

class CensoColumnasView(APIView):
    permission_classes = [IsGestion]

    def get(self, request):
        return Response({"columns": COLUMN_DEFS, "defaults": DEFAULT_COLUMNS})


class CensoListView(APIView):
    permission_classes = [IsGestion]

    def get(self, request):
        tenant = request.tenant
        if not getattr(tenant, "animales_enabled", False):
            return Response({"detail": "Módulo no habilitado."}, status=403)

        qs = _build_queryset(tenant, request.query_params)

        # Pagination
        try:
            page = max(1, int(request.query_params.get("page", 1)))
            page_size = min(200, max(1, int(request.query_params.get("page_size", 50))))
        except (ValueError, TypeError):
            page, page_size = 1, 50

        total = qs.count()
        start = (page - 1) * page_size
        animals = qs[start:start + page_size]

        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "results": [_serialize_animal(a) for a in animals],
        })


class CensoExportView(APIView):
    permission_classes = [IsGestion]

    def get(self, request):
        tenant = request.tenant
        if not getattr(tenant, "animales_enabled", False):
            return Response({"detail": "Módulo no habilitado."}, status=403)

        fmt = request.query_params.get("format", "excel")
        columns_param = request.query_params.get("columns", "")
        selected = [c.strip() for c in columns_param.split(",") if c.strip()] or DEFAULT_COLUMNS
        # Keep only valid keys
        valid_keys = {d["key"] for d in COLUMN_DEFS}
        selected = [c for c in selected if c in valid_keys]
        if not selected:
            selected = DEFAULT_COLUMNS

        col_labels = {d["key"]: d["label"] for d in COLUMN_DEFS}
        headers = [col_labels[k] for k in selected]

        qs = _build_queryset(tenant, request.query_params)
        raw_rows = [_serialize_animal(a) for a in qs]
        # Convert to list-of-lists for template compatibility
        rows = [[r.get(k, "") for k in selected] for r in raw_rows]

        if fmt == "pdf":
            return self._export_pdf(tenant, headers, rows)
        return self._export_excel(tenant, headers, selected, raw_rows)

    def _export_pdf(self, tenant, headers, rows):
        from django.template.loader import render_to_string
        from weasyprint import HTML  # type: ignore

        html_string = render_to_string("animals/censo_animales.html", {
            "tenant": tenant,
            "headers": headers,
            "rows": rows,
        })
        pdf_bytes = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="censo_animales.pdf"'
        return response

    def _export_excel(self, tenant, headers, selected, rows):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Animales"

        header_fill = PatternFill("solid", fgColor="051937")
        header_font = Font(bold=True, color="FFFFFF")

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(rows, start=2):
            for ci, key in enumerate(selected, start=1):
                ws.cell(row=ri, column=ci, value=row.get(key, ""))

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="censo_animales.xlsx"'
        return response
