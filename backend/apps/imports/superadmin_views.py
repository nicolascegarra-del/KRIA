import io
import uuid

import openpyxl
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.reports.storage import upload_bytes
from apps.tenants.models import Tenant
from core.permissions import IsSuperAdmin
from core.throttles import UploadRateThrottle
from .models import ImportJob
from .tasks import process_import_job, process_animal_import_job
from .views import ImportTemplateView, AnimalImportTemplateView


def _get_tenant_or_404(pk):
    try:
        return Tenant.objects.get(pk=pk), None
    except Tenant.DoesNotExist:
        return None, Response({"detail": "Asociación no encontrada."}, status=404)


# ── Templates ─────────────────────────────────────────────────────────────────

class SuperAdminImportTemplateSociosView(APIView):
    """GET /superadmin/importaciones/socios/template/"""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        return ImportTemplateView().get(request)


class SuperAdminImportTemplateAnimalesView(APIView):
    """GET /superadmin/importaciones/animales/template/"""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        return AnimalImportTemplateView().get(request)


# ── Socios import ──────────────────────────────────────────────────────────────

class SuperAdminImportValidateSociosView(APIView):
    """POST /superadmin/importaciones/<tenant_id>/socios/validate/"""
    permission_classes = [IsSuperAdmin]
    parser_classes = [MultiPartParser]
    throttle_classes = [UploadRateThrottle]

    def post(self, request, tenant_id):
        tenant, err = _get_tenant_or_404(tenant_id)
        if err:
            return err

        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            file_bytes = file.read()
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
        except Exception as e:
            return Response({"detail": f"No se pudo leer el archivo Excel: {str(e)}"}, status=400)

        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        advisory_cols = {"dni_nif", "nombre_razon_social", "email"}
        missing_cols = sorted(advisory_cols - set(headers))
        rows_preview = []
        advertencias = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            row_dict = {
                headers[i]: str(v if v is not None else "").strip()
                for i, v in enumerate(row) if i < len(headers)
            }
            if row_dict.get("dni_nif", "").upper().startswith("DNI"):
                continue
            dni = row_dict.get("dni_nif", "")
            email = row_dict.get("email", "")
            nombre = row_dict.get("nombre_razon_social", "")
            avisos = []
            if not dni:
                avisos.append("dni_nif vacío")
            if not email:
                avisos.append("email vacío")
            if not nombre:
                avisos.append("nombre_razon_social vacío")
            rows_preview.append({
                "fila": idx + 2, "dni_nif": dni,
                "email": email, "nombre_razon_social": nombre, "errores": avisos,
            })
            if avisos:
                advertencias.append({"fila": idx + 2, "errores": avisos})

        temp_key = f"imports/{tenant.slug}/validate/{uuid.uuid4()}/{file.name}"
        try:
            upload_bytes(temp_key, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            return Response({"detail": f"Error almacenando archivo: {str(e)}"}, status=500)

        total = len(rows_preview)
        return Response({
            "total_filas": total,
            "filas_ok": total - len(advertencias),
            "filas_con_error": len(advertencias),
            "errores": advertencias,
            "columnas_faltantes": missing_cols,
            "preview": rows_preview,
            "temp_key": temp_key,
        })


class SuperAdminImportConfirmSociosView(APIView):
    """POST /superadmin/importaciones/<tenant_id>/socios/confirm/"""
    permission_classes = [IsSuperAdmin]

    def post(self, request, tenant_id):
        tenant, err = _get_tenant_or_404(tenant_id)
        if err:
            return err

        temp_key = request.data.get("temp_key")
        if not temp_key:
            return Response({"detail": "temp_key es obligatorio."}, status=400)

        if not temp_key.startswith(f"imports/{tenant.slug}/"):
            return Response({"detail": "Invalid temp_key."}, status=400)

        job = ImportJob.objects.create(tenant=tenant, created_by=request.user, file_key=temp_key)
        process_import_job.delay(str(job.id))
        return Response({"job_id": str(job.id), "status": job.status}, status=202)


# ── Animales import ────────────────────────────────────────────────────────────

class SuperAdminImportValidateAnimalesView(APIView):
    """POST /superadmin/importaciones/<tenant_id>/animales/validate/"""
    permission_classes = [IsSuperAdmin]
    parser_classes = [MultiPartParser]
    throttle_classes = [UploadRateThrottle]

    def post(self, request, tenant_id):
        tenant, err = _get_tenant_or_404(tenant_id)
        if err:
            return err

        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            file_bytes = file.read()
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
        except Exception as e:
            return Response({"detail": f"No se pudo leer el archivo Excel: {str(e)}"}, status=400)

        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

        from apps.accounts.models import Socio as SocioModel
        rows_preview = []
        advertencias = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            row_dict = {
                headers[i]: str(v if v is not None else "").strip()
                for i, v in enumerate(row) if i < len(headers)
            }
            if (row_dict.get("numero_anilla", "").lower().startswith("número") or
                    row_dict.get("numero_anilla", "").lower().startswith("numero")):
                continue

            anilla = row_dict.get("numero_anilla", "")
            fecha_nac = row_dict.get("fecha_nacimiento", "")
            sexo = row_dict.get("sexo", "").upper()
            socio_dni = row_dict.get("socio_dni", "")
            socio_num = row_dict.get("socio_numero_socio", "")
            avisos = []

            if not anilla:
                avisos.append("numero_anilla vacío")
            if not fecha_nac:
                avisos.append("fecha_nacimiento vacía")
            if sexo not in ("M", "H", "MACHO", "HEMBRA"):
                avisos.append(f"sexo inválido ('{sexo}'), debe ser M o H")
            if not socio_dni and not socio_num:
                avisos.append("se necesita socio_dni o socio_numero_socio")
            else:
                if socio_dni:
                    if not SocioModel.all_objects.filter(tenant=tenant, dni_nif=socio_dni).exists():
                        avisos.append(f"socio con DNI '{socio_dni}' no encontrado en el sistema")
                else:
                    if not SocioModel.all_objects.filter(tenant=tenant, numero_socio=socio_num).exists():
                        avisos.append(f"socio con número '{socio_num}' no encontrado en el sistema")

            rows_preview.append({
                "fila": idx + 2, "numero_anilla": anilla,
                "fecha_nacimiento": fecha_nac, "sexo": sexo,
                "socio": socio_dni or socio_num or "—", "errores": avisos,
            })
            if avisos:
                advertencias.append({"fila": idx + 2, "errores": avisos})

        temp_key = f"imports/{tenant.slug}/animales/validate/{uuid.uuid4()}/{file.name}"
        try:
            upload_bytes(temp_key, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            return Response({"detail": f"Error almacenando archivo: {str(e)}"}, status=500)

        total = len(rows_preview)
        return Response({
            "total_filas": total,
            "filas_ok": total - len(advertencias),
            "filas_con_error": len(advertencias),
            "errores": advertencias,
            "columnas_faltantes": [],
            "preview": rows_preview,
            "temp_key": temp_key,
        })


class SuperAdminImportConfirmAnimalesView(APIView):
    """POST /superadmin/importaciones/<tenant_id>/animales/confirm/"""
    permission_classes = [IsSuperAdmin]

    def post(self, request, tenant_id):
        tenant, err = _get_tenant_or_404(tenant_id)
        if err:
            return err

        temp_key = request.data.get("temp_key")
        if not temp_key:
            return Response({"detail": "temp_key es obligatorio."}, status=400)

        if not temp_key.startswith(f"imports/{tenant.slug}/"):
            return Response({"detail": "Invalid temp_key."}, status=400)

        job = ImportJob.objects.create(tenant=tenant, created_by=request.user, file_key=temp_key)
        process_animal_import_job.delay(str(job.id))
        return Response({"job_id": str(job.id), "status": job.status}, status=202)


# ── Job status ─────────────────────────────────────────────────────────────────

class SuperAdminImportJobStatusView(APIView):
    """GET /superadmin/importaciones/job/<job_id>/"""
    permission_classes = [IsSuperAdmin]

    def get(self, request, job_id):
        try:
            job = ImportJob.objects.get(pk=job_id)
        except ImportJob.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        return Response({
            "id": str(job.id),
            "status": job.status,
            "result_summary": job.result_summary,
            "error_log": job.error_log,
            "created_at": job.created_at,
            "finished_at": job.finished_at,
        })
