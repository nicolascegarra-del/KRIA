"""
Sprint 7/8 — Tests de plantilla Excel y flujo de importación en 2 fases.

Cubre:
  - GET /imports/template/ devuelve un archivo Excel (.xlsx) con las cabeceras correctas.
  - POST /imports/validate/ con Excel válido → preview sin errores.
  - POST /imports/validate/ con columna requerida faltante → 400.
  - POST /imports/validate/ con filas con errores → preview muestra errores por fila.
  - POST /imports/confirm/ con temp_key → crea ImportJob (202).
  - Socio no puede acceder a /imports/template/ → 403.
"""
import io
import pytest
import openpyxl
from unittest.mock import patch


TEMPLATE_URL = "/api/v1/imports/template/"
VALIDATE_URL = "/api/v1/imports/validate/"
CONFIRM_URL = "/api/v1/imports/confirm/"


def _make_excel(columns, rows):
    """Helper: genera un Excel en memoria con las columnas y filas indicadas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.mark.django_db
class TestImportTemplate:

    def test_template_devuelve_xlsx(self, gestion_client, tenant):
        """GET /imports/template/ devuelve un archivo Excel válido con cabeceras."""
        resp = gestion_client.get(TEMPLATE_URL)
        assert resp.status_code == 200
        assert "spreadsheetml" in resp["Content-Type"]

        # Leer el Excel y verificar cabeceras
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "dni_nif" in headers
        assert "nombre_razon_social" in headers
        assert "email" in headers

    def test_template_tiene_fila_ejemplo(self, gestion_client, tenant):
        """La plantilla incluye una fila de ejemplo."""
        resp = gestion_client.get(TEMPLATE_URL)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 2  # cabecera + al menos una fila ejemplo

    def test_socio_no_puede_descargar_template(self, socio_client):
        """Socio recibe 403 al acceder a la plantilla."""
        resp = socio_client.get(TEMPLATE_URL)
        assert resp.status_code == 403

    def test_validate_excel_valido(self, gestion_client, tenant):
        """POST /imports/validate/ con Excel bien formado → preview sin errores críticos."""
        excel = _make_excel(
            ["dni_nif", "nombre_razon_social", "email", "first_name"],
            [["12345678Z", "Granja Test S.L.", "test@granja.es", "Juan"]],
        )
        with patch("apps.imports.views.upload_bytes", return_value="temp/key.xlsx"):
            resp = gestion_client.post(
                VALIDATE_URL,
                {"file": (excel, "socios.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                format="multipart",
            )
        assert resp.status_code == 200, resp.data
        assert resp.data["total_filas"] == 1
        assert resp.data["filas_con_error"] == 0
        assert "temp_key" in resp.data

    def test_validate_columna_requerida_faltante(self, gestion_client, tenant):
        """POST /imports/validate/ sin columna 'email' → 400."""
        excel = _make_excel(
            ["dni_nif", "nombre_razon_social"],  # sin email
            [["12345678Z", "Granja Test S.L."]],
        )
        resp = gestion_client.post(
            VALIDATE_URL,
            {"file": (excel, "socios.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            format="multipart",
        )
        assert resp.status_code == 400
        assert "email" in resp.data["detail"].lower()

    def test_validate_filas_con_errores(self, gestion_client, tenant):
        """POST /imports/validate/ con fila sin email → errores reportados por fila."""
        excel = _make_excel(
            ["dni_nif", "nombre_razon_social", "email"],
            [
                ["12345678Z", "Granja Buena S.L.", "ok@granja.es"],
                ["", "Granja Mala S.L.", ""],  # fila con errores
            ],
        )
        with patch("apps.imports.views.upload_bytes", return_value="temp/key.xlsx"):
            resp = gestion_client.post(
                VALIDATE_URL,
                {"file": (excel, "socios.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                format="multipart",
            )
        assert resp.status_code == 200
        assert resp.data["filas_con_error"] >= 1
        assert len(resp.data["errores"]) >= 1

    def test_confirm_crea_import_job(self, gestion_client, tenant):
        """POST /imports/confirm/ con temp_key válido → crea ImportJob (202)."""
        with patch("apps.imports.views.process_import_job.delay"):
            resp = gestion_client.post(
                CONFIRM_URL,
                {"temp_key": "imports/demo/validate/abc/socios.xlsx"},
                format="json",
            )
        assert resp.status_code == 202, resp.data
        assert "job_id" in resp.data

    def test_confirm_sin_temp_key_400(self, gestion_client, tenant):
        """POST /imports/confirm/ sin temp_key → 400."""
        resp = gestion_client.post(CONFIRM_URL, {}, format="json")
        assert resp.status_code == 400
