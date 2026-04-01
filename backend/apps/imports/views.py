import io
import uuid

import openpyxl
from django.http import HttpResponse
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.reports.storage import upload_bytes
from core.permissions import IsGestion
from core.throttles import UploadRateThrottle
from .models import ImportJob
from .tasks import process_import_job, process_animal_import_job

# Expected columns for the import template
IMPORT_COLUMNS = [
    "dni_nif", "nombre_razon_social", "email",
    "first_name", "last_name", "telefono",
    "domicilio", "municipio", "codigo_postal", "provincia",
    "numero_cuenta", "numero_socio", "codigo_rega",
    "fecha_alta", "cuota_anual_pagada", "estado", "razon_baja", "fecha_baja",
]

# Human-readable descriptions shown as a subtitle row in the template
COLUMN_DESCRIPTIONS = [
    "DNI / NIF / NIE / CIF",
    "Razón social o nombre completo  [OBLIGATORIO]",
    "Correo electrónico",
    "Nombre",
    "Apellidos",
    "Teléfono",
    "Domicilio / Calle",
    "Municipio",
    "Código postal",
    "Provincia",
    "IBAN / Número de cuenta",
    "Número de socio",
    "Código REGA",
    "Fecha de alta AAAA-MM-DD",
    "Año cuota pagada (ej: 2025)",
    "Estado: ALTA o BAJA  (por defecto ALTA)",
    "Razón de baja (solo si estado=BAJA)",
    "Fecha de baja AAAA-MM-DD (solo si estado=BAJA)",
]

EXAMPLE_ROW = [
    "12345678Z", "Granja Ejemplo S.L.", "socio@ejemplo.es",
    "Juan", "García López", "612000000",
    "Calle Mayor 1", "Madrid", "28001", "Madrid",
    "ES9121000418450200051332", "00001", "ES280101000001",
    "2020-03-15", "2025", "ALTA", "", "",
]


class SocioImportView(APIView):
    permission_classes = [IsGestion]
    throttle_classes = [UploadRateThrottle]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        tenant = request.tenant
        file_key = f"imports/{tenant.slug}/{uuid.uuid4()}/{file.name}"

        try:
            from apps.reports.storage import get_minio_client
            from django.conf import settings

            client = get_minio_client()
            client.put_object(
                settings.MINIO_BUCKET_NAME,
                file_key,
                file,
                length=file.size,
                content_type=file.content_type,
            )
        except Exception as e:
            return Response({"detail": f"Storage error: {str(e)}"}, status=500)

        job = ImportJob.objects.create(
            tenant=tenant,
            created_by=request.user,
            file_key=file_key,
        )

        process_import_job.delay(str(job.id))

        return Response({"job_id": str(job.id), "status": job.status}, status=202)


class ImportJobStatusView(APIView):
    permission_classes = [IsGestion]

    def get(self, request, job_id):
        try:
            job = ImportJob.objects.get(pk=job_id, tenant=request.tenant)
        except ImportJob.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        return Response({
            "id": str(job.id),
            "status": job.status,
            "result_summary": job.result_summary,
            "error_log": job.error_log,
            "created_at": job.created_at,
            "finished_at": job.finished_at,
        })


class ImportTemplateView(APIView):
    """GET /api/v1/imports/template/ — devuelve la plantilla Excel para importar socios."""
    permission_classes = [IsGestion]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Socios"

        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        desc_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        desc_font = Font(italic=True, color="555555", size=9)
        example_font = Font(italic=True, color="999999", size=9)

        # Row 1 — column keys (used by the import engine)
        for col, col_name in enumerate(IMPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Row 2 — human-readable descriptions
        for col, desc in enumerate(COLUMN_DESCRIPTIONS, 1):
            cell = ws.cell(row=2, column=col, value=desc)
            cell.fill = desc_fill
            cell.font = desc_font
            cell.alignment = Alignment(wrap_text=True)

        # Row 3 — example data
        for col, value in enumerate(EXAMPLE_ROW, 1):
            cell = ws.cell(row=3, column=col, value=value)
            cell.font = example_font

        ws.row_dimensions[2].height = 30

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 42)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="plantilla_socios.xlsx"'
        return response


class ImportValidateView(APIView):
    """
    POST /api/v1/imports/validate/ — fase 1: valida el Excel y devuelve preview de errores
    sin guardar nada en la base de datos. Guarda el archivo temp en MinIO.
    """
    permission_classes = [IsGestion]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            file_bytes = file.read()
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
        except Exception as e:
            return Response({"detail": f"No se pudo leer el archivo Excel: {str(e)}"}, status=400)

        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

        # Columnas que faltan en el archivo — advertencia informativa, no bloquea
        advisory_cols = {"dni_nif", "nombre_razon_social", "email"}
        missing_cols = sorted(advisory_cols - set(headers))

        rows_preview = []
        advertencias = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            # Ignorar filas completamente vacías
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            row_dict = {
                headers[i]: str(v if v is not None else "").strip()
                for i, v in enumerate(row)
                if i < len(headers)
            }
            # Skip template description/example header rows
            # (e.g. "DNI / NIF / NIE / CIF [OBLIGATORIO]")
            if row_dict.get("dni_nif", "").upper().startswith("DNI"):
                continue

            dni = row_dict.get("dni_nif", "")
            email = row_dict.get("email", "")
            nombre = row_dict.get("nombre_razon_social", "")
            avisos = []

            if not dni:
                avisos.append("dni_nif vacío")
            if not email:
                avisos.append("email vacío")
            if not nombre:
                avisos.append("nombre_razon_social vacío")

            rows_preview.append({
                "fila": idx + 2,
                "dni_nif": dni,
                "email": email,
                "nombre_razon_social": nombre,
                "errores": avisos,
            })
            if avisos:
                advertencias.append({"fila": idx + 2, "errores": avisos})

        # Save temp file to MinIO
        tenant = request.tenant
        temp_key = f"imports/{tenant.slug}/validate/{uuid.uuid4()}/{file.name}"
        try:
            upload_bytes(temp_key, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            return Response({"detail": f"Error almacenando archivo: {str(e)}"}, status=500)

        total = len(rows_preview)
        return Response({
            "total_filas": total,
            "filas_ok": total - len(advertencias),
            "filas_con_error": len(advertencias),
            "errores": advertencias,
            "columnas_faltantes": missing_cols,
            "preview": rows_preview,
            "temp_key": temp_key,
        })


class ImportConfirmView(APIView):
    """
    POST /api/v1/imports/confirm/ — fase 2: ejecuta la importación real con el temp_key.
    """
    permission_classes = [IsGestion]

    def post(self, request):
        temp_key = request.data.get("temp_key")
        if not temp_key:
            return Response({"detail": "temp_key es obligatorio."}, status=400)

        tenant = request.tenant

        # Security: ensure temp_key belongs to this tenant
        expected_prefix = f"imports/{tenant.slug}/"
        if not temp_key.startswith(expected_prefix):
            return Response({"detail": "Invalid temp_key."}, status=400)

        # Create ImportJob pointing at the already-uploaded temp file
        job = ImportJob.objects.create(
            tenant=tenant,
            created_by=request.user,
            file_key=temp_key,
        )

        process_import_job.delay(str(job.id))

        return Response({"job_id": str(job.id), "status": job.status}, status=202)


# ── Animal import ──────────────────────────────────────────────────────────────

ANIMAL_IMPORT_COLUMNS = [
    "numero_anilla", "fecha_nacimiento", "sexo", "variedad",
    "socio_dni", "socio_numero_socio",
    "padre_anilla", "madre_anilla", "madre_lote_externo",
    "fecha_incubacion", "ganaderia_nacimiento", "ganaderia_actual",
]

ANIMAL_COLUMN_DESCRIPTIONS = [
    "Número de anilla  [OBLIGATORIO]",
    "Fecha de nacimiento AAAA-MM-DD  [OBLIGATORIO]",
    "Sexo: M (macho) o H (hembra)  [OBLIGATORIO]",
    "Variedad: SALMON / PLATA / SIN_DEFINIR (defecto SALMON)",
    "DNI/NIF del socio propietario  [OBLIGATORIO si no hay numero_socio]",
    "Número de socio alternativo para localizar al socio",
    "Anilla del padre (debe existir en el sistema)",
    "Anilla de la madre (debe existir en el sistema)",
    "Descripción libre del lote externo de la madre",
    "Fecha de incubación AAAA-MM-DD",
    "Ganadería de nacimiento",
    "Ganadería actual",
]

ANIMAL_EXAMPLE_ROW = [
    "ES-1234-2023-A", "2023-04-15", "M", "SALMON",
    "12345678Z", "",
    "ES-0001-2020-A", "", "",
    "2023-03-20", "Granja Ejemplo", "Granja Ejemplo",
]


class AnimalImportTemplateView(APIView):
    """GET /api/v1/imports/animales/template/ — plantilla Excel para importar animales."""
    permission_classes = [IsGestion]

    def get(self, request):
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Animales"

        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        desc_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        desc_font = Font(italic=True, color="555555", size=9)
        example_font = Font(italic=True, color="999999", size=9)

        for col, col_name in enumerate(ANIMAL_IMPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col, desc in enumerate(ANIMAL_COLUMN_DESCRIPTIONS, 1):
            cell = ws.cell(row=2, column=col, value=desc)
            cell.fill = desc_fill
            cell.font = desc_font
            cell.alignment = Alignment(wrap_text=True)

        for col, value in enumerate(ANIMAL_EXAMPLE_ROW, 1):
            cell = ws.cell(row=3, column=col, value=value)
            cell.font = example_font

        ws.row_dimensions[2].height = 30
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 42)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="plantilla_animales.xlsx"'
        return response


class AnimalImportValidateView(APIView):
    """POST /api/v1/imports/animales/validate/ — fase 1: valida sin guardar."""
    permission_classes = [IsGestion]
    parser_classes = [MultiPartParser]

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file provided."}, status=400)

        try:
            file_bytes = file.read()
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws = wb.active
        except Exception as e:
            return Response({"detail": f"No se pudo leer el archivo Excel: {str(e)}"}, status=400)

        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

        rows_preview = []
        advertencias = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            row_dict = {
                headers[i]: str(v if v is not None else "").strip()
                for i, v in enumerate(row)
                if i < len(headers)
            }
            # Skip template description rows
            if row_dict.get("numero_anilla", "").lower().startswith("número") or \
               row_dict.get("numero_anilla", "").lower().startswith("numero"):
                continue

            anilla = row_dict.get("numero_anilla", "")
            fecha_nac = row_dict.get("fecha_nacimiento", "")
            sexo = row_dict.get("sexo", "").upper()
            socio_dni = row_dict.get("socio_dni", "")
            socio_num = row_dict.get("socio_numero_socio", "")
            avisos = []

            if not anilla:
                avisos.append("numero_anilla vacío")
            if not fecha_nac:
                avisos.append("fecha_nacimiento vacía")
            if sexo not in ("M", "H", "MACHO", "HEMBRA"):
                avisos.append(f"sexo inválido ('{sexo}'), debe ser M o H")
            if not socio_dni and not socio_num:
                avisos.append("se necesita socio_dni o socio_numero_socio")

            rows_preview.append({
                "fila": idx + 2,
                "numero_anilla": anilla,
                "fecha_nacimiento": fecha_nac,
                "sexo": sexo,
                "socio": socio_dni or socio_num or "—",
                "errores": avisos,
            })
            if avisos:
                advertencias.append({"fila": idx + 2, "errores": avisos})

        tenant = request.tenant
        temp_key = f"imports/{tenant.slug}/animales/validate/{uuid.uuid4()}/{file.name}"
        try:
            upload_bytes(temp_key, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            return Response({"detail": f"Error almacenando archivo: {str(e)}"}, status=500)

        total = len(rows_preview)
        return Response({
            "total_filas": total,
            "filas_ok": total - len(advertencias),
            "filas_con_error": len(advertencias),
            "errores": advertencias,
            "columnas_faltantes": [],
            "preview": rows_preview,
            "temp_key": temp_key,
        })


class AnimalImportConfirmView(APIView):
    """POST /api/v1/imports/animales/confirm/ — fase 2: ejecuta la importación real."""
    permission_classes = [IsGestion]

    def post(self, request):
        temp_key = request.data.get("temp_key")
        if not temp_key:
            return Response({"detail": "temp_key es obligatorio."}, status=400)

        tenant = request.tenant
        expected_prefix = f"imports/{tenant.slug}/"
        if not temp_key.startswith(expected_prefix):
            return Response({"detail": "Invalid temp_key."}, status=400)

        job = ImportJob.objects.create(
            tenant=tenant,
            created_by=request.user,
            file_key=temp_key,
        )

        process_animal_import_job.delay(str(job.id))

        return Response({"job_id": str(job.id), "status": job.status}, status=202)
