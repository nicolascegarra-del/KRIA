"""
Celery tasks for PDF and Excel report generation.
"""
import logging
import traceback
import uuid

from celery import shared_task
from django.utils import timezone

logger = logging.getLogger(__name__)


def _update_job(job, status, file_key="", error=""):
    job.status = status
    job.file_key = file_key
    job.error_log = error
    job.finished_at = timezone.now()
    job.save(update_fields=["status", "file_key", "error_log", "finished_at"])


@shared_task(name="reports.generate_report", queue="reports", bind=True, max_retries=0)
def generate_report(self, job_id: str):
    from .models import ReportJob
    try:
        job = ReportJob.all_objects.select_related("tenant", "created_by").get(pk=job_id)
    except ReportJob.DoesNotExist:
        return

    job.status = ReportJob.Status.PROCESSING
    job.save(update_fields=["status"])

    try:
        rt = job.report_type
        fmt = job.params.get("formato", "pdf")
        if rt == ReportJob.ReportType.INVENTORY:
            file_key = _gen_inventory_excel(job) if fmt == "excel" else _gen_inventory_pdf(job)
        elif rt == ReportJob.ReportType.INDIVIDUAL:
            file_key = _gen_individual_pdf(job)
        elif rt == ReportJob.ReportType.GENEALOGY_CERT:
            file_key = _gen_genealogy_cert(job)
        elif rt == ReportJob.ReportType.LIBRO_GENEALOGICO:
            file_key = _gen_libro_genealogico(job)
        elif rt == ReportJob.ReportType.CATALOGO_REPRODUCTORES:
            file_key = _gen_catalogo_reproductores_excel(job) if fmt == "excel" else _gen_catalogo_reproductores(job)
        elif rt == ReportJob.ReportType.AUDITORIA:
            file_key = _gen_auditoria_pdf(job)
        else:
            raise ValueError(f"Unknown report type: {rt}")

        _update_job(job, ReportJob.Status.DONE, file_key=file_key)

    except Exception:
        _update_job(job, ReportJob.Status.FAILED, error=traceback.format_exc())


# ── PDF generators ────────────────────────────────────────────────────────────

def _render_pdf(template_name: str, context: dict) -> bytes:
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS  # type: ignore
    html_string = render_to_string(template_name, context)
    return HTML(string=html_string).write_pdf()


def _anio(animal) -> str:
    """Return birth year as string, tolerating both fecha_nacimiento and missing."""
    try:
        return str(animal.fecha_nacimiento.year)
    except Exception:
        return ""


def _gen_inventory_pdf(job) -> str:
    from apps.animals.models import Animal

    socio_id = job.params.get("socio_id")
    socio_nombre = None
    if socio_id:
        animals = Animal.all_objects.filter(
            tenant=job.tenant, socio_id=socio_id
        ).select_related("socio", "padre", "madre_animal", "evaluacion").order_by("variedad", "numero_anilla")
        first = animals.first()
        if first:
            socio_nombre = first.socio.nombre_razon_social
    else:
        animals = Animal.all_objects.filter(
            tenant=job.tenant
        ).select_related("socio", "padre", "madre_animal", "evaluacion").order_by("variedad", "numero_anilla")

    context = {
        "tenant": job.tenant,
        "animals": animals,
        "socio_nombre": socio_nombre,
        "watermark": "COPIA",
    }
    pdf_bytes = _render_pdf("reports/inventory.html", context)
    key = f"reports/{job.tenant.slug}/inventory/{uuid.uuid4()}.pdf"
    from .storage import upload_bytes
    return upload_bytes(key, pdf_bytes, "application/pdf")


def _gen_individual_pdf(job) -> str:
    from apps.animals.models import Animal

    animal_id = job.params.get("animal_id")
    animal = Animal.all_objects.select_related(
        "socio",
        "padre", "padre__padre", "padre__madre_animal",
        "madre_animal", "madre_animal__padre", "madre_animal__madre_animal",
        "madre_lote", "evaluacion",
    ).get(pk=animal_id, tenant=job.tenant)

    padre = animal.padre
    madre = animal.madre_animal
    gen = {
        "padre": padre,
        "madre": madre,
        "abuelo_paterno": getattr(padre, "padre", None) if padre else None,
        "abuela_paterna": getattr(padre, "madre_animal", None) if padre else None,
        "abuelo_materno": getattr(madre, "padre", None) if madre else None,
        "abuela_materna": getattr(madre, "madre_animal", None) if madre else None,
    }
    has_genealogy = bool(padre or madre or animal.madre_lote_id or animal.madre_lote_externo)

    context = {"tenant": job.tenant, "animal": animal, "gen": gen, "has_genealogy": has_genealogy}
    pdf_bytes = _render_pdf("reports/individual.html", context)
    key = f"reports/{job.tenant.slug}/individual/{animal_id}.pdf"
    from .storage import upload_bytes
    return upload_bytes(key, pdf_bytes, "application/pdf")


def _gen_genealogy_cert(job) -> str:
    from apps.animals.models import Animal
    from apps.animals.serializers import _build_genealogy_node

    animal_id = job.params.get("animal_id")
    animal = Animal.all_objects.get(pk=animal_id, tenant=job.tenant)

    tree = _build_genealogy_node(animal)
    context = {"tenant": job.tenant, "animal": animal, "tree": tree, "watermark": "CERTIFICADO"}
    pdf_bytes = _render_pdf("reports/genealogy_cert.html", context)
    key = f"reports/{job.tenant.slug}/genealogy/{animal_id}.pdf"
    from .storage import upload_bytes
    return upload_bytes(key, pdf_bytes, "application/pdf")


def _gen_inventory_excel(job) -> str:
    """Generate Inventario de Animales as XLSX — optimized for working in Excel."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from apps.animals.models import Animal

    socio_id = job.params.get("socio_id")
    if socio_id:
        animals_qs = Animal.all_objects.filter(
            tenant=job.tenant, socio_id=socio_id
        ).select_related("socio", "padre", "madre_animal").prefetch_related("evaluacion")
    else:
        animals_qs = Animal.all_objects.filter(
            tenant=job.tenant
        ).select_related("socio", "padre", "madre_animal").prefetch_related("evaluacion").order_by("variedad", "numero_anilla")

    animals_list = list(animals_qs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    primary_hex = (job.tenant.primary_color or "#1565C0").lstrip("#")
    if len(primary_hex) == 3:
        primary_hex = "".join(c * 2 for c in primary_hex)

    hdr_fill  = PatternFill(start_color=primary_hex, end_color=primary_hex, fill_type="solid")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    alt_fill  = PatternFill(start_color="EFF3F8", end_color="EFF3F8", fill_type="solid")
    foot_fill = PatternFill(start_color="E2EAF4", end_color="E2EAF4", fill_type="solid")
    brd = Border(
        left=Side(style="thin", color="C8D0DA"),
        right=Side(style="thin", color="C8D0DA"),
        top=Side(style="thin", color="C8D0DA"),
        bottom=Side(style="thin", color="C8D0DA"),
    )

    HEADERS = [
        "Nº Anilla", "Año Nac.", "Sexo", "Variedad", "Estado",
        "Padre (Anilla)", "Madre (Anilla)",
        "Socio / Titular", "DNI / NIF", "Cód. REGA",
        "Punt. Media",
    ]
    N = len(HEADERS)
    last_col = get_column_letter(N)

    # ── Fila 1: Título ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"Inventario de Animales  ·  {job.tenant.name}"
    c.font = Font(bold=True, size=14, color=primary_hex)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill(start_color="F5F8FC", end_color="F5F8FC", fill_type="solid")
    ws.row_dimensions[1].height = 32

    # ── Fila 2: Cabeceras ───────────────────────────────────────────────────────
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
    ws.row_dimensions[2].height = 24

    # ── Filas de datos ──────────────────────────────────────────────────────────
    for i, animal in enumerate(animals_list):
        r = i + 3
        try:
            media = float(animal.evaluacion.puntuacion_media)
        except Exception:
            media = None
        row = [
            animal.numero_anilla,
            animal.fecha_nacimiento.year if animal.fecha_nacimiento else None,
            animal.get_sexo_display(),
            animal.get_variedad_display(),
            animal.get_estado_display(),
            animal.padre.numero_anilla if animal.padre else "",
            animal.madre_animal.numero_anilla if animal.madre_animal else "",
            animal.socio.nombre_razon_social,
            animal.socio.dni_nif,
            animal.socio.codigo_rega or "",
            media,
        ]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = brd
            if fill:
                c.fill = fill
            if col in (1, 6, 7):
                c.alignment = Alignment(horizontal="center")
            elif col == 2:
                c.alignment = Alignment(horizontal="center")
            elif col == 11:
                c.alignment = Alignment(horizontal="center")
                if val is not None:
                    c.number_format = "0.00"
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Fila de totales ─────────────────────────────────────────────────────────
    tr = len(animals_list) + 3
    ws.merge_cells(f"A{tr}:{last_col}{tr}")
    c = ws[f"A{tr}"]
    c.value = f"Total: {len(animals_list)} animales"
    c.font = Font(bold=True, size=9, italic=True, color="445566")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.fill = foot_fill
    ws.row_dimensions[tr].height = 18

    # ── Auto-filtro, freeze, anchos ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.freeze_panes = "A3"
    for i, w in enumerate([16, 9, 9, 13, 15, 16, 16, 30, 14, 14, 11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    key = f"reports/{job.tenant.slug}/inventory/{uuid.uuid4()}.xlsx"
    from .storage import upload_bytes
    return upload_bytes(key, buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _gen_catalogo_reproductores_excel(job) -> str:
    """Generate Catálogo de Reproductores as XLSX — optimized for working in Excel."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from apps.animals.models import Animal

    animals_qs = Animal.all_objects.filter(
        tenant=job.tenant, reproductor_aprobado=True,
    ).select_related("socio").prefetch_related("evaluacion").order_by("variedad", "numero_anilla")
    animals_list = list(animals_qs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo Reproductores"

    primary_hex = (job.tenant.primary_color or "#1565C0").lstrip("#")
    if len(primary_hex) == 3:
        primary_hex = "".join(c * 2 for c in primary_hex)

    hdr_fill  = PatternFill(start_color=primary_hex, end_color=primary_hex, fill_type="solid")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    alt_fill  = PatternFill(start_color="EFF3F8", end_color="EFF3F8", fill_type="solid")
    foot_fill = PatternFill(start_color="E2EAF4", end_color="E2EAF4", fill_type="solid")
    brd = Border(
        left=Side(style="thin", color="C8D0DA"),
        right=Side(style="thin", color="C8D0DA"),
        top=Side(style="thin", color="C8D0DA"),
        bottom=Side(style="thin", color="C8D0DA"),
    )

    HEADERS = [
        "Nº Anilla", "Año Nac.", "Sexo", "Variedad",
        "Socio / Titular", "DNI / NIF", "Cód. REGA",
        "Cabeza", "Cola", "Pecho/Abd.", "Muslos/Tarsos", "Cresta/Babilla", "Color",
        "Punt. Media",
        "Ganadería Nac.", "Fecha Nac.",
    ]
    N = len(HEADERS)
    last_col = get_column_letter(N)

    # ── Fila 1: Título ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"Catálogo de Reproductores  ·  {job.tenant.name}"
    c.font = Font(bold=True, size=14, color=primary_hex)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill(start_color="F5F8FC", end_color="F5F8FC", fill_type="solid")
    ws.row_dimensions[1].height = 32

    # ── Fila 2: Cabeceras ───────────────────────────────────────────────────────
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
    ws.row_dimensions[2].height = 24

    # ── Filas de datos ──────────────────────────────────────────────────────────
    for i, animal in enumerate(animals_list):
        r = i + 3
        ev = getattr(animal, "evaluacion", None)
        try:
            ev_real = ev if ev and ev.pk else None
        except Exception:
            ev_real = None
        row = [
            animal.numero_anilla,
            animal.fecha_nacimiento.year if animal.fecha_nacimiento else None,
            animal.get_sexo_display(),
            animal.get_variedad_display(),
            animal.socio.nombre_razon_social,
            animal.socio.dni_nif,
            animal.socio.codigo_rega or "",
            ev_real.cabeza if ev_real else None,
            ev_real.cola if ev_real else None,
            ev_real.pecho_abdomen if ev_real else None,
            ev_real.muslos_tarsos if ev_real else None,
            ev_real.cresta_babilla if ev_real else None,
            ev_real.color if ev_real else None,
            float(ev_real.puntuacion_media) if ev_real else None,
            animal.ganaderia_nacimiento or "",
            animal.fecha_nacimiento.strftime("%d/%m/%Y") if animal.fecha_nacimiento else "",
        ]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = brd
            if fill:
                c.fill = fill
            if col == 1:
                c.alignment = Alignment(horizontal="center")
            elif col in (2, 8, 9, 10, 11, 12, 13):
                c.alignment = Alignment(horizontal="center")
            elif col == 14:
                c.alignment = Alignment(horizontal="center")
                if val is not None:
                    c.number_format = "0.00"
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Fila de totales ─────────────────────────────────────────────────────────
    tr = len(animals_list) + 3
    ws.merge_cells(f"A{tr}:{last_col}{tr}")
    c = ws[f"A{tr}"]
    c.value = f"Total reproductores aprobados: {len(animals_list)}"
    c.font = Font(bold=True, size=9, italic=True, color="445566")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.fill = foot_fill
    ws.row_dimensions[tr].height = 18

    # ── Auto-filtro, freeze, anchos ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.freeze_panes = "A3"
    for i, w in enumerate([16, 9, 9, 13, 28, 14, 14, 8, 8, 10, 12, 12, 8, 10, 22, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    key = f"reports/{job.tenant.slug}/catalogo/{uuid.uuid4()}.xlsx"
    from .storage import upload_bytes
    return upload_bytes(key, buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _gen_libro_genealogico(job) -> str:
    """
    Generate ARCA/Ministerio format Excel Libro Genealógico — optimized for working in Excel.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from apps.animals.models import Animal
    from django.utils import timezone

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro Genealógico"

    PRIMARY = "1565C0"
    hdr_fill  = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    alt_fill  = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    foot_fill = PatternFill(start_color="D6E8F7", end_color="D6E8F7", fill_type="solid")
    brd = Border(
        left=Side(style="thin", color="B8C8D8"),
        right=Side(style="thin", color="B8C8D8"),
        top=Side(style="thin", color="B8C8D8"),
        bottom=Side(style="thin", color="B8C8D8"),
    )

    HEADERS = [
        "Nº Anilla", "Año Nac.", "Fecha Nac.", "Sexo", "Variedad",
        "Padre (Anilla)", "Madre (Anilla)", "Lote Madre",
        "Fecha Incubación", "Ganadería Nacimiento",
        "Socio / Titular", "DNI / NIF", "Cód. REGA",
        "Estado", "Punt. Media",
    ]
    N = len(HEADERS)
    last_col = get_column_letter(N)

    # ── Fila 1: Título ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"Libro Genealógico  ·  {job.tenant.name}"
    c.font = Font(bold=True, size=14, color=PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    ws.row_dimensions[1].height = 32

    # ── Fila 2: Subtítulo ───────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = f"Formato ARCA / Ministerio  ·  Generado el {timezone.now().strftime('%d/%m/%Y a las %H:%M')}h"
    c.font = Font(italic=True, size=8, color="7090A8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    ws.row_dimensions[2].height = 16

    # ── Fila 3: Cabeceras ───────────────────────────────────────────────────────
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
    ws.row_dimensions[3].height = 24

    # ── Datos ───────────────────────────────────────────────────────────────────
    animals = Animal.all_objects.filter(tenant=job.tenant).select_related(
        "socio", "padre", "madre_animal", "madre_lote"
    ).prefetch_related("evaluacion").order_by("variedad", "numero_anilla")

    animals_list = list(animals)
    for i, animal in enumerate(animals_list):
        r = i + 4
        try:
            media = float(animal.evaluacion.puntuacion_media)
        except Exception:
            media = None

        lote_madre = ""
        if animal.madre_lote_id:
            try:
                lote_madre = animal.madre_lote.nombre
            except Exception:
                pass
        if not lote_madre and animal.madre_lote_externo:
            lote_madre = f"[Ext.] {animal.madre_lote_externo}"

        row = [
            animal.numero_anilla,
            animal.fecha_nacimiento.year if animal.fecha_nacimiento else None,
            animal.fecha_nacimiento.strftime("%d/%m/%Y") if animal.fecha_nacimiento else "",
            animal.get_sexo_display(),
            animal.get_variedad_display(),
            animal.padre.numero_anilla if animal.padre else "",
            animal.madre_animal.numero_anilla if animal.madre_animal else "",
            lote_madre,
            animal.fecha_incubacion.strftime("%d/%m/%Y") if animal.fecha_incubacion else "",
            animal.ganaderia_nacimiento or "",
            animal.socio.nombre_razon_social,
            animal.socio.dni_nif,
            animal.socio.codigo_rega or "",
            animal.get_estado_display(),
            media,
        ]
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = brd
            if fill:
                c.fill = fill
            if col in (1, 6, 7):
                c.alignment = Alignment(horizontal="center")
            elif col == 2:
                c.alignment = Alignment(horizontal="center")
            elif col == 15:
                c.alignment = Alignment(horizontal="center")
                if val is not None:
                    c.number_format = "0.00"
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Fila de totales ─────────────────────────────────────────────────────────
    tr = len(animals_list) + 4
    ws.merge_cells(f"A{tr}:{last_col}{tr}")
    c = ws[f"A{tr}"]
    c.value = f"Total registros: {len(animals_list)}"
    c.font = Font(bold=True, size=9, italic=True, color="336688")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.fill = foot_fill
    ws.row_dimensions[tr].height = 18

    # ── Auto-filtro, freeze, anchos ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:{last_col}3"
    ws.freeze_panes = "A4"
    for i, w in enumerate([16, 9, 13, 9, 13, 16, 16, 22, 14, 26, 30, 14, 14, 14, 11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    key = f"reports/{job.tenant.slug}/libro-genealogico/{uuid.uuid4()}.xlsx"
    from .storage import upload_bytes
    return upload_bytes(key, buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _gen_catalogo_reproductores(job) -> str:
    """
    Generate Catálogo Reproductores PDF: 1 page per animal, with radar chart.
    """
    from apps.animals.models import Animal
    import io, base64

    animals = Animal.all_objects.filter(
        tenant=job.tenant,
        reproductor_aprobado=True,
    ).select_related("socio").prefetch_related("evaluacion").order_by("variedad", "numero_anilla")

    pages = []
    for animal in animals:
        radar_b64 = ""
        try:
            ev = animal.evaluacion
            radar_b64 = _generate_radar_chart(ev)
        except Exception:
            pass
        pages.append({"animal": animal, "radar_b64": radar_b64})

    context = {"tenant": job.tenant, "pages": pages}
    pdf_bytes = _render_pdf("reports/catalogo_reproductores.html", context)
    key = f"reports/{job.tenant.slug}/catalogo/{uuid.uuid4()}.pdf"
    from .storage import upload_bytes
    return upload_bytes(key, pdf_bytes, "application/pdf")


def _gen_auditoria_pdf(job) -> str:
    from apps.audits.models import AuditoriaSession, CriterioEvaluacion

    auditoria_id = job.params.get("auditoria_id")
    auditoria = AuditoriaSession.all_objects.select_related(
        "socio", "tenant"
    ).prefetch_related(
        "animales_evaluados__animal",
        "respuestas_instalacion__pregunta",
    ).get(pk=auditoria_id, tenant=job.tenant)

    criterios = list(CriterioEvaluacion.objects.filter(tenant=job.tenant, is_active=True).order_by("orden", "nombre"))

    # Pre-process animal scores for easy template rendering
    animales_detalle = []
    for av in auditoria.animales_evaluados.all():
        scores = []
        for c in criterios:
            scores.append({
                "nombre": c.nombre,
                "multiplicador": c.multiplicador,
                "valor": av.puntuaciones.get(str(c.id), 0),
            })
        pct = None
        if av.puntuacion_maxima and av.puntuacion_maxima > 0:
            pct = round(float(av.puntuacion_total / av.puntuacion_maxima * 100), 1)
        animales_detalle.append({
            "anilla": av.animal.numero_anilla if av.animal else (av.numero_anilla_manual or "—"),
            "puntuacion_total": av.puntuacion_total,
            "puntuacion_maxima": av.puntuacion_maxima,
            "porcentaje": pct,
            "notas": av.notas,
            "scores": scores,
        })

    context = {
        "tenant": job.tenant,
        "auditoria": auditoria,
        "criterios": criterios,
        "animales_detalle": animales_detalle,
    }
    pdf_bytes = _render_pdf("reports/auditoria.html", context)
    key = f"reports/{job.tenant.slug}/auditorias/{auditoria_id}.pdf"
    from .storage import upload_bytes
    return upload_bytes(key, pdf_bytes, "application/pdf")


def _generate_radar_chart(evaluacion) -> str:
    """Generate a base64-encoded radar/spider chart PNG using matplotlib."""
    import io, base64
    import numpy as np
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    categories = ["Cabeza", "Cola", "Pecho/Abd.", "Muslos/Tarsos", "Cresta/Babilla", "Color"]
    values = [
        evaluacion.cabeza, evaluacion.cola, evaluacion.pecho_abdomen,
        evaluacion.muslos_tarsos, evaluacion.cresta_babilla, evaluacion.color,
    ]

    N = len(categories)
    angles = [n / float(N) * 2 * np.pi for n in range(N)]
    angles += angles[:1]
    values_plot = values + values[:1]

    fig, ax = plt.subplots(figsize=(4, 4), subplot_kw=dict(polar=True))
    ax.plot(angles, values_plot, "o-", linewidth=2, color="#1565C0")
    ax.fill(angles, values_plot, alpha=0.25, color="#1565C0")
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(categories, size=9)
    ax.set_ylim(0, 10)
    ax.set_yticks([2, 4, 6, 8, 10])
    ax.set_yticklabels(["2", "4", "6", "8", "10"], size=7)
    ax.set_title(f"Puntuación media: {evaluacion.puntuacion_media}", size=10, pad=15)

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode("utf-8")
