"""
Vistas del módulo TABLAS.
Todas las operaciones están restringidas a gestión (IsGestion).
"""
import io
from django.db.models import Count
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import IsGestion
from .models import TablaControl, TablaColumna, TablaEntrada
from .serializers import (
    TablaControlListSerializer,
    TablaControlDetailSerializer,
    TablaControlWriteSerializer,
    TablaEntradaSerializer,
)

# ── Campos del socio disponibles para mostrar en tablas ──────────────────────
SOCIO_FIELD_LABELS = {
    "numero_socio": "Nº Socio",
    "nombre_razon_social": "Nombre / Razón Social",
    "dni_nif": "DNI / NIF",
    "email": "Email",
    "telefono": "Teléfono",
    "municipio": "Municipio",
    "provincia": "Provincia",
    "estado": "Estado",
    "fecha_alta": "Fecha de Alta",
    "cuota_anual_pagada": "Cuota Pagada (año)",
}


def _check_tablas_enabled(request):
    if not getattr(request.tenant, "tablas_enabled", False):
        return Response({"detail": "El módulo Tablas no está habilitado."}, status=403)
    return None


# ── CRUD de Tablas ────────────────────────────────────────────────────────────

class TablaControlListCreateView(APIView):
    permission_classes = [IsGestion]

    def get(self, request):
        err = _check_tablas_enabled(request)
        if err:
            return err
        qs = TablaControl.objects.filter(tenant=request.tenant).annotate(
            columnas_count=Count("columnas")
        )
        return Response(TablaControlListSerializer(qs, many=True).data)

    def post(self, request):
        err = _check_tablas_enabled(request)
        if err:
            return err
        ser = TablaControlWriteSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        tabla = ser.save(tenant=request.tenant)
        return Response(TablaControlDetailSerializer(tabla).data, status=201)


class TablaControlDetailView(APIView):
    permission_classes = [IsGestion]

    def _get_tabla(self, request, pk):
        try:
            return TablaControl.objects.get(pk=pk, tenant=request.tenant)
        except TablaControl.DoesNotExist:
            return None

    def get(self, request, pk):
        err = _check_tablas_enabled(request)
        if err:
            return err
        tabla = self._get_tabla(request, pk)
        if not tabla:
            return Response({"detail": "No encontrado."}, status=404)
        return Response(TablaControlDetailSerializer(tabla).data)

    def put(self, request, pk):
        err = _check_tablas_enabled(request)
        if err:
            return err
        tabla = self._get_tabla(request, pk)
        if not tabla:
            return Response({"detail": "No encontrado."}, status=404)
        ser = TablaControlWriteSerializer(tabla, data=request.data)
        ser.is_valid(raise_exception=True)
        tabla = ser.save()
        return Response(TablaControlDetailSerializer(tabla).data)

    def patch(self, request, pk):
        err = _check_tablas_enabled(request)
        if err:
            return err
        tabla = self._get_tabla(request, pk)
        if not tabla:
            return Response({"detail": "No encontrado."}, status=404)
        ser = TablaControlWriteSerializer(tabla, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        tabla = ser.save()
        return Response(TablaControlDetailSerializer(tabla).data)

    def delete(self, request, pk):
        err = _check_tablas_enabled(request)
        if err:
            return err
        tabla = self._get_tabla(request, pk)
        if not tabla:
            return Response({"detail": "No encontrado."}, status=404)
        tabla.delete()
        return Response(status=204)


# ── Filas (entradas) de una tabla ────────────────────────────────────────────

class TablaFilasView(APIView):
    permission_classes = [IsGestion]

    def get(self, request, pk):
        err = _check_tablas_enabled(request)
        if err:
            return err
        try:
            tabla = TablaControl.objects.get(pk=pk, tenant=request.tenant)
        except TablaControl.DoesNotExist:
            return Response({"detail": "No encontrado."}, status=404)

        qs = TablaEntrada.objects.filter(tabla=tabla).select_related("socio", "socio__user").order_by(
            "socio__numero_socio"
        )

        # Search filter
        search = request.query_params.get("search", "").strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(
                Q(socio__nombre_razon_social__icontains=search)
                | Q(socio__numero_socio__icontains=search)
                | Q(socio__dni_nif__icontains=search)
                | Q(socio__user__email__icontains=search)
            )

        # Estado filter
        estado = request.query_params.get("estado", "").strip()
        if estado in ("ALTA", "BAJA"):
            qs = qs.filter(socio__estado=estado)

        # Ordering
        ordering = request.query_params.get("ordering", "numero_socio").strip()
        ALLOWED_ORDERING = {
            "numero_socio": "socio__numero_socio",
            "-numero_socio": "-socio__numero_socio",
            "nombre": "socio__nombre_razon_social",
            "-nombre": "-socio__nombre_razon_social",
            "estado": "socio__estado",
            "-estado": "-socio__estado",
        }
        db_ordering = ALLOWED_ORDERING.get(ordering, "socio__numero_socio")
        qs = qs.order_by(db_ordering)

        return Response(TablaEntradaSerializer(qs, many=True).data)


# ── Actualización inline de una entrada ──────────────────────────────────────

class TablaEntradaUpdateView(APIView):
    permission_classes = [IsGestion]

    def patch(self, request, pk, socio_id):
        err = _check_tablas_enabled(request)
        if err:
            return err
        try:
            tabla = TablaControl.objects.get(pk=pk, tenant=request.tenant)
        except TablaControl.DoesNotExist:
            return Response({"detail": "No encontrado."}, status=404)
        try:
            entrada = TablaEntrada.objects.get(tabla=tabla, socio_id=socio_id)
        except TablaEntrada.DoesNotExist:
            return Response({"detail": "Entrada no encontrada."}, status=404)

        nuevos_valores = request.data.get("valores", {})
        if not isinstance(nuevos_valores, dict):
            return Response({"detail": "Se esperaba un objeto en 'valores'."}, status=400)

        # Validate that all keys are valid column IDs for this table
        col_ids = set(str(c.id) for c in tabla.columnas.all())
        invalid = set(nuevos_valores.keys()) - col_ids
        if invalid:
            return Response({"detail": f"Columnas inválidas: {invalid}"}, status=400)

        entrada.valores.update(nuevos_valores)
        entrada.save(update_fields=["valores", "updated_at"])
        return Response(TablaEntradaSerializer(entrada).data)


# ── Sincronización de socios (añadir nuevos socios a entradas existentes) ─────

class TablaSyncSociosView(APIView):
    """Añade entradas para socios que aún no estén en la tabla."""
    permission_classes = [IsGestion]

    def post(self, request, pk):
        err = _check_tablas_enabled(request)
        if err:
            return err
        try:
            tabla = TablaControl.objects.get(pk=pk, tenant=request.tenant)
        except TablaControl.DoesNotExist:
            return Response({"detail": "No encontrado."}, status=404)

        from apps.accounts.models import Socio
        socios_con_entrada = set(
            TablaEntrada.objects.filter(tabla=tabla).values_list("socio_id", flat=True)
        )
        socios_nuevos = Socio.objects.filter(tenant=request.tenant).exclude(id__in=socios_con_entrada)
        nuevas = TablaEntrada.objects.bulk_create([
            TablaEntrada(tabla=tabla, socio=s, valores={})
            for s in socios_nuevos
        ])
        return Response({"added": len(nuevas)})


# ── Exportación ──────────────────────────────────────────────────────────────

class TablaExportView(APIView):
    """Exporta la tabla a PDF o Excel de forma síncrona."""
    permission_classes = [IsGestion]

    def post(self, request, pk):
        err = _check_tablas_enabled(request)
        if err:
            return err
        try:
            tabla = TablaControl.objects.prefetch_related("columnas").get(
                pk=pk, tenant=request.tenant
            )
        except TablaControl.DoesNotExist:
            return Response({"detail": "No encontrado."}, status=404)

        formato = request.data.get("formato", "pdf").lower()
        if formato == "excel":
            return self._export_excel(tabla)
        return self._export_pdf(tabla, request.tenant)

    def _get_entradas(self, tabla):
        return (
            TablaEntrada.objects
            .filter(tabla=tabla)
            .select_related("socio", "socio__user")
            .order_by("socio__numero_socio")
        )

    def _export_pdf(self, tabla, tenant):
        from django.template.loader import render_to_string
        from weasyprint import HTML  # type: ignore

        columnas = list(tabla.columnas.order_by("orden", "nombre"))
        entradas = self._get_entradas(tabla)

        rows = []
        for e in entradas:
            socio_data = {
                "numero_socio": e.socio.numero_socio,
                "nombre_razon_social": e.socio.nombre_razon_social,
                "dni_nif": e.socio.dni_nif,
                "email": e.socio.user.email if e.socio.user_id else "",
                "telefono": e.socio.telefono,
                "municipio": e.socio.municipio,
                "provincia": e.socio.provincia,
                "estado": e.socio.estado,
                "fecha_alta": str(e.socio.fecha_alta) if e.socio.fecha_alta else "",
                "cuota_anual_pagada": str(e.socio.cuota_anual_pagada) if e.socio.cuota_anual_pagada else "",
            }
            # Build aligned list for template (avoids dynamic key access in Django templates)
            socio_values = [socio_data.get(field, "") for field in tabla.socio_columns]
            col_values = []
            for col in columnas:
                val = e.valores.get(str(col.id), "")
                if col.tipo == "CHECKBOX":
                    val = "Sí" if val is True or val == "true" else "No"
                col_values.append(val)
            rows.append({
                "socio_values": socio_values,
                "col_values": col_values,
            })

        socio_headers = [SOCIO_FIELD_LABELS.get(f, f) for f in tabla.socio_columns]
        html_string = render_to_string("tablas/tabla_control.html", {
            "tenant": tenant,
            "tabla": tabla,
            "socio_headers": socio_headers,
            "columnas": columnas,
            "rows": rows,
        })
        pdf_bytes = HTML(string=html_string).write_pdf()

        filename = f"tabla_{tabla.nombre[:40].replace(' ', '_')}.pdf"
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _export_excel(self, tabla):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        columnas = list(tabla.columnas.order_by("orden", "nombre"))
        entradas = self._get_entradas(tabla)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tabla.nombre[:31]

        header_fill = PatternFill("solid", fgColor="051937")
        header_font = Font(bold=True, color="FFFFFF")

        # Build header row
        headers = []
        for field in tabla.socio_columns:
            headers.append(SOCIO_FIELD_LABELS.get(field, field))
        for col in columnas:
            headers.append(col.nombre)

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for ri, e in enumerate(entradas, start=2):
            socio_data = {
                "numero_socio": e.socio.numero_socio,
                "nombre_razon_social": e.socio.nombre_razon_social,
                "dni_nif": e.socio.dni_nif,
                "email": e.socio.user.email if e.socio.user_id else "",
                "telefono": e.socio.telefono,
                "municipio": e.socio.municipio,
                "provincia": e.socio.provincia,
                "estado": e.socio.estado,
                "fecha_alta": str(e.socio.fecha_alta) if e.socio.fecha_alta else "",
                "cuota_anual_pagada": e.socio.cuota_anual_pagada or "",
            }
            ci = 1
            for field in tabla.socio_columns:
                ws.cell(row=ri, column=ci, value=str(socio_data.get(field, "")))
                ci += 1
            for col in columnas:
                val = e.valores.get(str(col.id), "")
                if col.tipo == "CHECKBOX":
                    val = "Sí" if val is True or val == "true" else "No"
                ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
                ci += 1

        # Auto column width
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"tabla_{tabla.nombre[:40].replace(' ', '_')}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


# ── Campos disponibles del socio ─────────────────────────────────────────────

class SocioFieldsView(APIView):
    permission_classes = [IsGestion]

    def get(self, request):
        return Response([
            {"key": k, "label": v}
            for k, v in SOCIO_FIELD_LABELS.items()
        ])
